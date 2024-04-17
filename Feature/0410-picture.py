"""
通过指定规则匹配证书图片网址 , 下载图片到本地并写入Excel文件指定位置
"""
import PIL
import io
import os
import requests
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as xlImage

quality_certificate_number = input("输入质检证书所在列号:")
master_bar_code = input("输入主条形码所在列号:")
picture_fill = input("输入需要填充的图片所在列号:")
path_to_fill = input("输入要补充的表格所在的文件夹路径(路径连接用[/]):")
path_to_save = input("输入图片保存本地的文件夹路径:")
authorization = input("输入更新后的Authorization:")

def get_images_from_folder(folder_path, save_path):
    for file_name in os.listdir(folder_path):
        print(file_name)
        if not file_name.endswith('.xlsx'):
            continue
        file_path = os.path.join(folder_path, file_name)
        wb = load_workbook(filename=file_path)
        ws = wb.active
        image_paths = {}
        for row in ws.iter_rows():
            # if row[0].row <= 1:
            if row[0].row <= 5:
                continue
            cert_no = ''
            for cell in row:
                if cell.column_letter != quality_certificate_number:
                    continue
                if cell.value == '质检号' or cell.value == '总计':
                    continue
                cert_no = cell.value
                print(cert_no)
            for cell in row:
                if cell.column_letter != master_bar_code:
                    continue
                if cell.value == '主条形码' or cell.value == '总计':
                    continue
                url_tiaoma = f'http://web.eastshow.cn/ytx/store/bar_code/{cell.value}'
                print(url_tiaoma)
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36 Edg/114.0.1823.51',
                    'Accept': '*/*',
                    'Content-Type': 'application/json;charset=UTF-8',
                    'Authorization': authorization,
                    'Referer': 'http://web.eastshow.cn/member/list/detail/119378'
                }
                response = requests.get(url_tiaoma, headers=headers)

                # 调试语句
                # print(response.status_code)  # 打印响应状态码
                # print(response.content)  # 打印响应内容

                data = response.json().get('data')
                if data is None or 'fiimage' not in data:
                    print("Image not found for barcode:", cell.value)
                    continue

                image_content = data['fiimage']

                if image_content == "https://externalresource.oss-cn-beijing.aliyuncs.com/store-order/8bda14af35fd414c9064b89151494fbd.jpg":
                    if cert_no is None:
                        image_content = "https://externalresource.oss-cn-beijing.aliyuncs.com/store-order/8bda14af35fd414c9064b89151494fbd.jpg?x-oss-process=image/resize,m_fixed,h_100,w_100"
                    elif cert_no.startswith(('PZ', 'Q', 'A')):
                        year = cert_no[2:4] if cert_no.startswith('PZ') else cert_no[1:3]
                        month = cert_no[4:6] if cert_no.startswith('PZ') else cert_no[3:5]
                        image_content = f"http://jpg.gtc-china.cn:8081/PIC/PZ/20{year}/{month}/{cert_no}.jpg" if cert_no.startswith('PZ') \
                            else f"http://2010.zpsx.cn:8080/GICCUG/GZ/20{year}/{month}/{cert_no}.jpg" if cert_no.startswith(('Q', 'A')) \
                            else "https://externalresource.oss-cn-beijing.aliyuncs.com/store-order/8bda14af35fd414c9064b89151494fbd.jpg?x-oss-process=image/resize,m_fixed,h_100,w_100"
                    else:
                        image_content = "https://externalresource.oss-cn-beijing.aliyuncs.com/store-order/8bda14af35fd414c9064b89151494fbd.jpg?x-oss-process=image/resize,m_fixed,h_100,w_100"
                else:
                    image_content += '?x-oss-process=image/resize,m_fixed,h_100,w_100'

                img_data = requests.get(image_content)
                img_path = os.path.join(save_path, f"{cell.value}.jpg")
                with open(img_path, "wb") as img_file:
                    img_file.write(img_data.content)

                image = PIL.Image.open(img_path)
                if image.mode == "RGBA":
                    image = image.convert("RGB")
                resized_image = image.resize((100,100))
                resized_image.save(img_path)

                image_paths[cell.value] = img_path

        for barcode, img_path in image_paths.items():
            img = xlImage(img_path)
            img.width = 100
            img.height = 100
            for row in ws.iter_rows():
                # if row[0].row <= 1:
                if row[0].row <= 5:
                    continue
                cell_barcode = ''
                cell_img = ''
                barcode_cell_position = ''
                img_cell_position = ''
                for cell in row:
                    if cell.column_letter == master_bar_code and cell.value != '总计' and cell.value != '主条形码':
                        cell_barcode = cell.value
                        barcode_cell_position = cell
                    if cell.column_letter == master_bar_code and cell.value == '总计':
                        continue
                    if cell.column_letter == picture_fill:
                        cell_img = cell.value
                        img_cell_position = cell
                if cell_barcode == barcode:
                    ws.add_image(img, img_cell_position.coordinate)
                    break

        wb.save(file_path)

    print('程序结束 , 请到文件夹查看')

get_images_from_folder(path_to_fill, path_to_save)
