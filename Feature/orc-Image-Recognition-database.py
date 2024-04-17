"""
    需求 :  获取mysql数据库的数据 , 表中有一字段是付款凭证的图片网址 , 识别图片内容 , 并sql根据查询将表格数据和识别的内容输出到excel文件中
    1 . 连接mysql数据库,根据查询获取到表的数据 ;
    2 . Python中的 Pillow 库来读取图片 , 以及使用 OpenCV 或者 PyTesseract 来进行图片内容的识别 ,
    Image.open() 函数只能打开本地文件，无法直接打开远程 URL。要处理远程图片，可以使用 requests 库来下载图片后再进行识别 ;
    3 . 通过 requests.get(image_url) 方法下载图片，然后使用 io.BytesIO(image_data) 将其读取为二进制数据，最后通过 Image.open() 方法将二进制数据转换为图像对象。
    (这种方式避免了将图片下载到本地硬盘上，而是直接在内存中处理图像数据，这对于简单的图像识别任务是足够的，并且更加高效
    识别图片中的文字 , 并将识别到的内容存储到 Excel中) ;
    4 . 判断识别出来的内容 , 去除空格和换行符\n , 截取需要的文本导出到Excel中

"""

import pymysql
import requests
from PIL import Image
import io
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from datetime import datetime
import re


# 连接数据库
def connect_to_database():
    connection = pymysql.connect(
        host='8.135.43.39',
        user='information',
        password='information123',
        database='ytxEbao',
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )
    return connection


# 从数据库中获取数据并识别图片内容
def process_database_data():
    connection = connect_to_database()
    try:
        with connection.cursor() as cursor:
            # 修改SQL查询语句为你需要的内容
            sql = r"""SELECT 
                    woopv.create_at AS 创建时间 ,
                    woopv.id ,
                    CAST(woopv.order_id AS CHAR) AS order_id , 
                    CASE 
                        WHEN woopv.pay_type = 9 THEN '扫呗支付'
                        WHEN woopv.pay_type = 5 THEN '通联支付'
                        WHEN woopv.pay_type = 1 THEN '微信支付'
                        WHEN woopv.pay_type = 2 THEN '支付宝支付'
                        WHEN woopv.pay_type = 3 THEN '付呗支付宝支付'
                        WHEN woopv.pay_type = 4 THEN '付呗微信支付'
                        WHEN woopv.pay_type = 6 THEN '富掌柜支付'
                        WHEN woopv.pay_type = 7 THEN '银行卡转账'
                        WHEN woopv.pay_type = 8 THEN '积分'
                        ELSE '未分类'
                    END AS 支付方式 , 
                    woopv.pay_amount AS 付款金额 ,
                    woopv.voucher_img_path AS 图片网址 
                    FROM `weChat_order_offline_payment_voucher` AS woopv
                    WHERE woopv.create_at >= '2024-04-01 00:00:00'
                    AND woopv.create_at <= '2024-04-10 23:59:59'"""
            cursor.execute(sql)
            result = cursor.fetchall()

            # 初始化Excel工作簿和表格
            wb = Workbook()
            ws = wb.active
            ws.append(['创建时间', 'id', 'oeder_id', '支付方式', '付款金额', '图片网址', 'Content'])

            # 创建日期格式和文本格式
            date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD HH:MM:SS')
            text_style = NamedStyle(name='text_style')

            for row in result:
                voucher_img_path = row['图片网址']

                # 读取图片并识别内容
                content = recognize_image_content(voucher_img_path)

                # 判断识别到的内容
                content = process_recognized_content(content)

                # 将数据写入Excel表格
                ws.append(
                    [row['创建时间'], row['id'], row['order_id'], row['支付方式'], row['付款金额'], row['图片网址'],
                     content])

            # 设置单元格格式
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
                for cell in row:
                    cell.style = date_style if cell.column == 1 else text_style

            # 保存Excel文件
            excel_file = r'D:\Users\Desktop\test\code\test_code\Feature\final_output.xlsx'
            wb.save(excel_file)
            print("Data saved to", excel_file)

    finally:
        connection.close()


# 识别图片内容
def recognize_image_content(image_url):
    try:
        # 下载图片
        response = requests.get(image_url)
        response.raise_for_status()  # 检查是否下载成功

        # 将下载的内容读取为二进制数据
        image_data = response.content

        # 将二进制数据转换为内存中的图像
        img = Image.open(io.BytesIO(image_data))

        # 使用Tesseract识别图像内容
        content = pytesseract.image_to_string(img, lang='chi_sim')

        # 去除空格和换行符
        content = content.replace(' ', '').replace('\n', '')

        # 去除所有中英文状态下的符号
        content = re.sub(r'[^\w\s]', '', content)
        return content

    except Exception as e:
        print("Error occurred while processing image:", e)
        return None


# 处理识别内容
def process_recognized_content(content):
    if content is None:
        return None

    # 判断识别内容中是否包含"订单尾号"和"商户单号"字段
    if "订单尾号" in content:
        index = content.find("订单尾号")
        return content[index + 4:index + 8]
    elif "商户单号" in content:
        index = content.find("商户单号")
        return content[index + 4:index + 22]
    else:
        return ""


# 主函数
def main():
    process_database_data()


if __name__ == "__main__":
    main()
