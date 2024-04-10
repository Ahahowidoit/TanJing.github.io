"""
        用于遍历指定路径下的Excel文件,并通过单元格内容与指定路径下的图片名称匹配 , 在表格的的指定位置插入图片
"""

import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as xlImage


excel_folder = input("请输入要插入图片的表格所在路径:")
image_folder = input("请输入图片所在的路径:")

def insert_images_to_excel(excel_folder, image_folder):
    # 遍历目标文件夹下的所有.xlsx文件
    for file_name in os.listdir(excel_folder):
        if not file_name.endswith('.xlsx'):
            continue
        excel_file_path = os.path.join(excel_folder, file_name)
        wb = load_workbook(filename=excel_file_path)
        ws = wb.active

        # 获取条码所在列的列号
        barcode_column = input("请输入条码所在的列号（例如：D）：")

        # 遍历Excel中的每一行
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            # ws.iter_rows工作表对象（Worksheet）的一个方法调用，用于在指定的行范围内迭代获取单元格
            # min_row = 2 指定了迭代的起始行为第二行（因为通常第一行是标题行）
            # max_col = ws.max_column 指定了迭代的最大列数为工作表的最大列数；
            # max_row = ws.max_row 指定了迭代的最大行数为工作表的最大行数。

            # 遍历每一行中的每一个单元格
            for cell in row:
                # cell.column_letter是 openpyxl 中单元格对象的一个属性，它返回单元格所在列的字母表示形式。
                # 例如，如果一个单元格位于 Excel 表格的第一列，那么它的 column_letter 属性将返回 'A'；
                # 如果位于第二列，则返回 'B'，以此类推
                if cell.column_letter == barcode_column:

                    # cell.value是openpyxl中单元格对象的一个属性，它返回该单元格中存储的值 ,该代码中用于获取每个单元格中存储的条码值
                    barcode_value = cell.value
                    if barcode_value is None:
                        continue

                    # 根据条码值在图片文件夹中查找对应的图片文件
                    image_file_path = os.path.join(image_folder, f"{barcode_value}.jpg")
                    if not os.path.exists(image_file_path):
                        continue

                    # 插入图片到Excel文件中
                    # 将图片文件的路径，传递给xlImage函数，表示要插入到Excel中的图片
                    img = xlImage(image_file_path)
                    img.width = 100
                    img.height = 100
                    # target_column_index = ws.max_column + 1  # 在最后一列的后面插入图片列
                    target_column_index = 5  # 在E列插入图片,E列对应的列索引是5

                    ws.column_dimensions[get_column_letter(target_column_index)].width = 15  # 设置图片列的宽度
                    ws.add_image(img, f"{get_column_letter(target_column_index)}{cell.row}")

        # 保存修改后的Excel文件
        wb.save(excel_file_path)


if __name__ == "__main__":
    # excel_folder = "E:/other/picture/0410picture-test"
    # image_folder = "E:/other/save_picture/0410picture"
    insert_images_to_excel(excel_folder, image_folder)
