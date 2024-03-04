# from PIL import Image
import PIL
import io
import os
import requests
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

quality_certificate_number = input("输入质检证书所在列号:")
master_bar_code = input("输入主条形码所在列号:")
picture_fill = input("输入需要填充的图片所在列号:")
path_to_fill = input("输入要补充的表格所在的文件夹路径(路径连接用[/]:")
path_to_save = input("输入图片保存本地的文件夹路径:")
authorization = input("输入更新后的Authorization:")

def get_images_from_folder(folder_path, save_path):
    i = 1
    for file_name in os.listdir(folder_path):
        print(file_name)
        if not file_name.endswith('.xlsx'):
            continue
        file_path = os.path.join(folder_path, file_name)
        wb = load_workbook(filename=file_path)
        ws = wb.active
        image_paths = {}
        for row in ws.iter_rows():
            if row[0].row <= 1:
            # if row[0].row <= 5:
                continue
            cert_no = ''
            for cell in row:
                if cell.column_letter != quality_certificate_number:
                    continue
                if cell.value == '质检号' or cell.value == '总计' :
                    continue
                cert_no = cell.value
                print(cert_no)
            for cell in row:
                if cell.column_letter != master_bar_code:
                    continue
                if cell.value == '主条形码' or cell.value == '总计'   :
                    continue
                # if cell.value is None :
                #     break
                url_tiaoma = f'http://web.eastshow.cn/ytx/store/bar_code/{cell.value}'
                if cell.value == '31171502':
                    pass
                print(url_tiaoma)
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36 Edg/114.0.1823.51',
                    'Accept': '*/*',
                    'Content-Type': 'application/json;charset=UTF-8',
                    # 'Authorization': 'YTX_STORE;eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJyb2xlIjoiMTEwIiwidW5pcXVlX25hbWUiOiLosK3pnZkiLCJ1c2VySWQiOiIxMDA1MjMiLCJpc3MiOiJyZXN0YXBpdXNlciIsImF1ZCI6IndZRGd2RkV3dUVIWkNySmxsdTFxSWlpRlVWcmcyeXA5IiwiZXhwIjoxNjk4OTk5ODU4LCJuYmYiOjE2OTg4MjcwNTh9.7NMosrHlqdPNSHbuB7H_X0XlbSnlCoEauiKd9F22vnw',
                    'Authorization': authorization ,
                    'Referer': 'http://web.eastshow.cn/member/list/detail/119378'
                    # 'Cookie': 'sessionid=xxxxxxxxxxxxxxx'
                }
                response = requests.get(url_tiaoma, headers=headers)
                image_content = json.loads(response.content)['data']['fiimage']
                url1 = image_content
                # url1 = "https://externalresource.oss-cn-beijing.aliyuncs.com/store-order/8bda14af35fd414c9064b89151494fbd.jpg"
                # cert_no = "PZ23032100202"

                if url1 == "https://externalresource.oss-cn-beijing.aliyuncs.com/store-order/8bda14af35fd414c9064b89151494fbd.jpg":
                    if cert_no is None:
                        url1 = "https://externalresource.oss-cn-beijing.aliyuncs.com/store-order/8bda14af35fd414c9064b89151494fbd.jpg?x-oss-process=image/resize,m_fixed,h_100,w_100"
                        img_data = requests.get(url1)
                    if cert_no is not None:
                        if cert_no[0:2] == 'PZ':
                            year = cert_no[2:4]
                            month = cert_no[4:6]
                            url1 = f"http://jpg.gtc-china.cn:8081/PIC/PZ/20{year}/{month}/{cert_no}.jpg"
                            img_data = requests.get(url1)
                            if img_data.status_code == 404:
                                img_data = requests.get("https://externalresource.oss-cn-beijing.aliyuncs.com/store-order/8bda14af35fd414c9064b89151494fbd.jpg?x-oss-process=image/resize,m_fixed,h_100,w_100")
                            # pass
                        elif cert_no[0] == 'Q' or cert_no[0] == 'A':
                            year = cert_no[1:3]
                            month = cert_no[3:5]
                            url1 = f"http://2010.zpsx.cn:8080/GICCUG/GZ/20{year}/{month}/{cert_no}.jpg"
                            img_data = requests.get(url1)
                            if img_data.status_code == 404:
                                url1 = f"http://2010.zpsx.cn:8080/GICCUG/GZ2/20{year}/{month}/{cert_no}.jpg"
                                img_data = requests.get(url1)
                            if img_data.status_code == 404:
                                url1 = "https://externalresource.oss-cn-beijing.aliyuncs.com/store-order/8bda14af35fd414c9064b89151494fbd.jpg?x-oss-process=image/resize,m_fixed,h_100,w_100"
                                img_data = requests.get(url1)
                        else:
                            url1 = "https://externalresource.oss-cn-beijing.aliyuncs.com/store-order/8bda14af35fd414c9064b89151494fbd.jpg?x-oss-process=image/resize,m_fixed,h_100,w_100"
                            img_data = requests.get(url1)
                                # pass
                else :
                    img_data = requests.get(url1 + '?x-oss-process=image/resize,m_fixed,h_100,w_100')

                # img_data = requests.get(url1+'?x-oss-process=image/resize,m_fixed,h_100,w_100')
                # img_data = requests.get(url1)
                img_path = os.path.join(save_path, f"{cell.value}.jpg")
                with open(img_path, "wb") as img_file:
                    img_file.write(img_data.content)
                image = PIL.Image.open(img_path)
                if image.mode == "RGBA":
                    image = image.convert("RGB")
                resized_image = image.resize((100,100))
                resized_image.save(img_path)
                image_paths[cell.value] = img_path
                i = i +1
        for barcode, img_path in image_paths.items():
            img = Image(img_path)
            img.wigth = 100
            img.height = 100
            for row in ws.iter_rows():
                if row[0].row <= 1:
                # if row[0].row <= 5:
                    continue
                cell_barcode = ''
                cell_img = ''
                barcode_cell_position = ''
                img_cell_position = ''
                for cell in row:
                    if cell.column_letter == master_bar_code and cell.value != '总计' and cell.value != '主条形码':
                        cell_barcode = cell.value
                        barcode_cell_position =cell
                    if cell.column_letter == master_bar_code and cell.value == '总计':
                        continue
                    if cell.column_letter == picture_fill:
                        cell_img = cell.value
                        img_cell_position = cell
                if cell_barcode == barcode:
                    ws.add_image(img, img_cell_position.coordinate)
                    break

        wb.save(file_path)

    print('程序结束 , 请到文件夹查看')
# 第一个参数为:要补充的表格所在的文件夹路径
# 第二个参数为:图片存储到本地的路径
# get_images_from_folder('E:/other/testpicture3', 'E:/other/save')
# get_images_from_folder('E:/other/picture/picture06', 'E:/other/save_picture/picture0625')
# get_images_from_folder('E:/other/picture/picture06/2306picture', 'E:/other/picture/picture06/picture06save')
# get_images_from_folder('E:/other/picture/1101picture', 'E:/other/save_picture/picture1101')
get_images_from_folder(path_to_fill ,path_to_save )


