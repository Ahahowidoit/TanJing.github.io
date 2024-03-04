"""
        此代码用于在Excel表格中填充图片
       根据文件中图片名称与表格中指定列的内容匹配 , 在Excel指定列中插入对应的图片
"""

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os

# 用户输入电子表格文件名、存放图片的文件夹路径和主条形码的索引
excel_filename = input("请输入电子表格文件名（包括路径和扩展名）：")
image_folder = input("请输入存放图片的文件夹路径：")
picture_row = input("请输入图片需要插入到的列标(用大写字母表示)：")
barcode_col_index = int(input("请输入表示图片名称所在列的索引(索引从1开始)："))
save_path="D:/Users/Desktop/test/code/picture_all/save_file"


# 加载用户指定的 Excel 文件
wb = load_workbook(filename=excel_filename)
ws = wb.active

# 循环处理每一行（从第二行开始）
for row in range(2, ws.max_row + 1):
    barcode = ws.cell(row, barcode_col_index).value  # 获取主条形码值
    image_name = f"{barcode}.jpg"
    image_path = os.path.join(image_folder, image_name)  # 拼接图片文件路径

    # 检查图片文件是否存在
    if os.path.exists(image_path):
        # cell = f'M{row}'  # 获取主条形码所在单元格位置
        cell = f'{picture_row}{row}'  # 获取主条形码所在单元格位置
        img = Image(image_path)
        ws.add_image(img, cell)  # 插入图片到对应的单元格
    else:
        print(f"图片文件不存在: {image_path}")

# 保存修改后的 Excel 文件
output_filename = input("请输入修改后的表格名称:")+'.xlsx'
print(f"修改后的文件保存为: {output_filename}")
wb.save(output_filename)

print(f"程序结束!\n请到:{save_path}路径下查看名为{output_filename}的输出文件")
